import openpyxl as pxl
import numpy as np
import matplotlib.pyplot as plt
from  colors_utils import *
from tqdm import trange
import pickle
import os
from datetime import datetime, timezone
from tkinter import filedialog
import re
import pandas as pd
import math


def save_spectral_dict_to_csv(data_dict, file_path='./IlluminationData.csv'):
    """
    Saves a dictionary of arrays to a CSV file.
    """
    df = pd.DataFrame(data_dict)
    
    df.to_csv(file_path, index=False)
    print(f"Data successfully saved to {file_path}")
    
def load_spectral_dict_from_csv(file_path):
    """
    Loads a CSV file and converts it back into a dictionary of NumPy arrays.
    """
    if not os.path.exists(file_path):
        print("File not found.")
        return {}

    df = pd.read_csv(file_path)
    data_dict = {column: df[column].values for column in df.columns}
    
    return data_dict


def load_correction_from_txt(file_path='./last_correction.txt'):
    if not os.path.exists(file_path):
        return {}
    correction = {}
    with open(file_path, 'r') as file:
        content = file.read()
        matches = re.findall(r"([\w\d_-]+) LED:\s+([\d\.]+|None)", content)
        for color, value in matches:
            correction[color] = None if value == "None" else float(value)
    return correction

def get_corrections(selected_LEDs=['Violet', 'Blue', 'Green', 'Yellow', 'Red']):
    """
    Returns a dictionary with updated values, or existing ones from the file 
    if no input is provided. New LEDs default to None.
    """
    print('\n--- CORRECTION PHASE ---')
    
    # 1. Load the current state from the file
    current_stored = load_correction_from_txt('./last_correction.txt')
    correction = {}
    
    for color in selected_LEDs:
        # Get existing value from file
        stored_val = current_stored.get(color, None)
        
        prompt = f"Power of {color} LED at 5V (Current: {stored_val} mW): "
        temp_input = input(prompt).strip()
        
        if temp_input != "":
            # Scenario: User typed a new value
            try:
                correction[color] = float(temp_input)
                print(f"  -> {color} updated to: {correction[color]} mW")
            except ValueError:
                print(f"  -> Invalid input. Using previous: {stored_val}")
                correction[color] = stored_val
        else:
            # Scenario: Empty input (Enter)
            # Use the stored value (could be 3.0, 5.0, or None if it's new)
            correction[color] = stored_val
            
    return correction

def save_correction_to_txt(final_correction, file_path='./last_correction.txt'):
    """
    Saves the merged dictionary directly to the file.
    """
    # We still merge just in case the file contains LEDs NOT in the current 'all_LEDs' list
    full_data = load_correction_from_txt(file_path)
    full_data.update(final_correction)

    with open(file_path, 'w') as f:
        f.write("Correction Data\n================\n\nPower Values (mW at 5V):\n")
        for color in sorted(full_data.keys()):
            val = full_data[color]
            val_str = "None" if val is None else f"{val}"
            f.write(f"{color} LED:\t{val_str}\tmW\n")
            
            


def get_latest_sheet_name(path):
    """Identifie le nom de la feuille avec la date YYYYMMDD la plus récente."""
    wb = pxl.load_workbook(path, read_only=True)
    dates = []
    for name in wb.sheetnames:
        try:
            dates.append(datetime.strptime(name, "%Y%m%d"))
        except ValueError:
            continue
    if not dates:
        raise ValueError("Aucune feuille au format YYYYMMDD trouvée dans le fichier Excel.")
    return max(dates).strftime("%Y%m%d")



def _find_led_columns(ws, selected_LEDs):
    """
    Parcourt la première ligne pour mapper les noms de LEDs aux indices de colonnes.
    Lève une erreur si une LED demandée est introuvable.
    """
    # On récupère les en-têtes (Ligne 1)
    headers = [str(ws.cell(row=1, column=c).value).strip() for c in range(1, ws.max_column + 1)]
    col_map = {}
    for led_name in selected_LEDs:
        found_idx = None
        for i, h in enumerate(headers):
            if led_name == h: # Recherche avec un valeur exact pour le match
                found_idx = i + 1 # openpyxl est base 1
                break
        
        if found_idx is None:
            raise ValueError(f"Erreur : La LED '{led_name}' est demandée mais n'a pas été trouvée dans les en-têtes du fichier Excel (En-têtes trouvés : {[h for h in headers if h != 'None']})")
        
        col_map[led_name] = found_idx
    return col_map

def _plot_results_with_spectra(results, spectra, sheet_name):
    """
    Génère une grille de graphiques : 
    Colonne 1 : Spectre d'émission (nm) | Colonne 2 : Courbe de puissance (V)
    """
    leds = [k for k in results.keys() if k not in ['voltages', 'ordered_leds']]
    n_leds = len(leds)
    
    # On crée une figure avec N lignes (une par LED) et 2 colonnes
    fig, axes = plt.subplots(n_leds, 2, figsize=(14, 4 * n_leds), squeeze=False)
    
    # Couleurs par défaut pour les spectres
    color_map = {
        "blue": "blue", "green": "green", "red": "red", 
        "white": "grey", "violet": "purple", "yellow": "orange"
    }

    for i, led in enumerate(leds):
        # --- AXE GAUCHE : SPECTRE ---
        ax_spec = axes[i, 0]
        x_spec = spectra.get("x_axis")
        
        # Recherche du spectre correspondant (match partiel)
        spec_key = next((k for k in spectra.keys() if led.lower() in k.lower() and k != "x_axis"), None)
        
        if spec_key and x_spec is not None:
            y_spec = spectra[spec_key]
            # Normalisation 0-1
            y_norm = np.clip(y_spec / np.max(y_spec), 0, None)
            
            # Choix de la couleur
            c = next((color_map[k] for k in color_map if k in led.lower()), "black")
            
            ax_spec.plot(x_spec, y_norm, color=c, lw=2)
            ax_spec.fill_between(x_spec, y_norm, color=c, alpha=0.2)
            ax_spec.set_title(f"Spectre d'émission : {led}")
            ax_spec.set_xlabel("Wavelength (nm)")
            ax_spec.set_ylabel("Normalized Intensity")
        else:
            ax_spec.text(0.5, 0.5, "Spectre non trouvé", ha='center')

        ax_spec.grid(True, alpha=0.3)

        # --- AXE DROIT : CALIBRATION ---
        ax_calib = axes[i, 1]
        ax_calib.plot(results['voltages'], results[led], marker='o', color='black', label='Calibration')
        ax_calib.set_title(f"Puissance Surfacique : {led}")
        ax_calib.set_xlabel("Tension (V)")
        ax_calib.set_ylabel("µW/cm²")
        ax_calib.grid(True, alpha=0.3)

    plt.suptitle(f"Analyse de Calibration : {sheet_name}", fontsize=16, fontweight='bold')
    plt.tight_layout(rect=[0, 0.03, 1, 0.97])
    plt.show()

def charge_calibration(path_excel, correction,path_spectra= './IlluminationData.pkl', selected_LEDs=['Violet', 'Blue', 'Green', 'Yellow', 'Red'], verbose=True):
    
    """
    Version étendue chargeant la calibration et les spectres.
    """
    # 1. Chargement des spectres (via ta fonction ledSpectrums ou direct)
    # On force plot=False ici car on va gérer le plot nous-mêmes dans la grille
    spectra_data = ledSpectrums(path=path_spectra, plot=False)
    
    # 2. Sélection de la feuille et ouverture Excel
    sheet_name = get_latest_sheet_name(path_excel)
    wb = pxl.load_workbook(path_excel, data_only=True)
    ws = wb[sheet_name]
    
    col_map = _find_led_columns(ws, selected_LEDs)
    voltages = np.array([ws.cell(row=i, column=1).value for i in range(4, 21)], dtype=float)
    
    results = {'voltages': voltages}
    
    # 3. Traitement par LED
    for led in selected_LEDs:
        col = col_map[led]
        p_curve = np.array([ws.cell(row=i, column=col).value for i in range(4, 21)], dtype=float)
        
        p_5v_excel = ws.cell(row=20, column=col).value
        ps_5v_excel = ws.cell(row=23, column=col).value
        
        ratio = ps_5v_excel / p_5v_excel if p_5v_excel and p_5v_excel != 0 else 0
        
        target_p_5v = correction.get(led)
        if target_p_5v is None:
            target_p_5v = p_5v_excel
            
        scaling = target_p_5v / p_5v_excel if p_5v_excel != 0 else 0
        results[led] = (p_curve * scaling) * ratio

    # 4. Visualisation combinée
    if verbose:
        _plot_results_with_spectra(results, spectra_data, sheet_name)
        
    return results


def get_voltages(Ptot, calibration, selected_LEDs = ['Violet', 'Blue', 'Green', 'Yellow', 'Red'], verbose=False):

    voltages = calibration['voltages']
    driving_tension = []
    
    for col_i in range(len(selected_LEDs)) :
        col_name = selected_LEDs[col_i]
        temp_P = float(Ptot[col_i])
                        
        # Return 0 voltage if the power is 0
        if temp_P == 0:
            driving_tension.append(0)
            if verbose:
                print(f"{col_name}: Power is 0, so voltage is 0 V.")
            continue

        
        calibration_list = calibration[col_name]
        if temp_P>calibration_list[-1]:
            raise ValueError('The given power value is too high. Try putting a lower value.')
    
        if temp_P in calibration_list :
            index=np.where(calibration_list==temp_P)
            res_voltage=voltages[index][0]
        
        else : 
            low_ind=0
            high_ind=len(calibration_list)-1
            
            for i in range(0,len(calibration_list)):
                if calibration_list[i]<temp_P and calibration_list[i]>calibration_list[low_ind]:
                    # print(calibration_list[i])
                    low_ind=i
                if calibration_list[i]>temp_P and calibration_list[i]<calibration_list[high_ind]:
                    high_ind=i
                #print(voltages[low_ind],voltages[high_ind])
            
            res_voltage = voltages[low_ind] + ( voltages[high_ind] - voltages[low_ind] ) * ( temp_P - calibration_list[low_ind] ) / ( calibration_list[high_ind] - calibration_list[low_ind] )
            driving_tension.append(res_voltage)
            
            if verbose:
                print('{}   :    {} V'.format(col_name, res_voltage) )
            
    driving_tension = np.array(driving_tension)
    return driving_tension