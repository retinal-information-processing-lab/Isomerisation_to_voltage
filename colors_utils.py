import numpy as np
import itertools
import matplotlib.pyplot as plt
from matplotlib.pyplot import *
import pickle
import os
import io
import plotly.graph_objects as go
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
import plotly.io as pio
import scipy
from tqdm.auto import tqdm
from datetime import datetime, timezone, timedelta
import matplotlib.pyplot as plt
import openpyxl as pxl



def prSpectrums(path="./PhotoReceptorData.pkl", plot=True):
    """
    Loads and optionally plots the spectral sensitivity of photoreceptors.

    This function retrieves spectral sensitivity data for various photoreceptors from a specified file
    and optionally plots their sensitivity curves.

    Parameters:
    - path (str, optional): Path to the photoreceptor data file. Default is "./PhotoReceptorData.pkl".
    - plot (bool, optional): If True, plots the spectral sensitivity curves. Default is True.

    Returns:
    - dict: A dictionary containing spectral sensitivity data with wavelength values under "x_axis"
      and sensitivities for each photoreceptor.

    Notes:
    - The photoreceptors considered are "Scones", "Mcones", "RedOpsin", "Rods", and "Mela".
    - The plot uses dashed lines with distinct colors and an alpha transparency of 0.5.
    - The function assumes the data file is in a NumPy-compatible format with `allow_pickle=True`.
    """

    
    prSpectrums = np.load(path, allow_pickle=True)
    colors={"Scones": "blue", "Mcones": "green", "RedOpsin": "red", "Rods": "black", "Mela": "purple"}
    if plot:
        x_axis = prSpectrums["x_axis"]
        for pr in prSpectrums.keys():
            if pr == "x_axis":
                continue    
            plt.plot(x_axis, (prSpectrums[pr]), label=pr, color=colors[pr], linestyle="--", alpha = 0.5)
        plt.legend()
        plt.title("PhotoReceptors Spectrums")
        plt.xlabel("Wavelength (nm)")
        plt.ylabel("Sensitivity")
    return prSpectrums



def ledSpectrums(path="./IlluminationData.pkl", plot=True):
    """
    Loads and optionally plots the spectral distributions of LED sources.

    This function retrieves spectral intensity data for various LEDs from a specified file
    and optionally plots their normalized emission spectra.

    Parameters:
    - path (str, optional): Path to the LED spectral data file. Default is "./IlluminationData.pkl".
    - plot (bool, optional): If True, plots the normalized spectral intensity curves. Default is True.

    Returns:
    - dict: A dictionary containing spectral intensity data with wavelength values under "x_axis"
      and emission spectra for each LED.

    Notes:
    - The LEDs considered are "Blue", "Green", "Red", "White", "Violet", and "Yellow".
    - Emission spectra are normalized by their maximum value and clipped at zero to avoid negative values.
    - The function assumes the data file is in a NumPy-compatible format with `allow_pickle=True`.
    """

    ledSpectrum = np.load(path, allow_pickle=True)
    colors = {"Blue": "blue", "Green": "green", "Red": "red", "White":"grey", "Violet": "purple", "Yellow": "orange"}

    if plot:
        x_axis = ledSpectrum["x_axis"]
        for led in ledSpectrum.keys():
            if led in ["x_axis","help"]:
                continue    
                
            if led not in colors.keys():
                colors[led] = get_led_color(ledSpectrum[led], ledSpectrum[x_axis])
            plt.plot(x_axis, (np.clip(ledSpectrum[led]/max(ledSpectrum[led]), 0, None)), label=led, color=colors[led], alpha = 0.7)
        plt.legend()
        plt.title("LED Spectrums")
        plt.xlabel("Wavelength (nm)")
        plt.ylabel("Intensity")
    return ledSpectrum


def get_isomerisation_matrix(selected_opsins, selected_leds, opsinDATA_path="./PhotoReceptorData.pkl", ledDATA_path="./IlluminationData.pkl", acDATA={"Scones": 0.2, "Mcones": 0.2, "RedOpsin": 0.002, "Rods": 0.5, "Mela": 0.2}, x_axis="x_axis"):
    
    """
    Computes the isomerisation matrix for a given set of opsins and LED spectra.

    This function calculates the isomerisation rates of selected opsins when exposed to a given set of LEDs,
    based on spectral data stored in specified files. This is the conversion matrix not the actual isomerrisation computation.

    Parameters:
    - selected_opsins (list): List of opsin names corresponding to keys in the opsin data file.
    - selected_leds (list): List of LED names corresponding to keys in the LED data file.
    - opsinDATA_path (str, optional): Path to the opsin spectral data file. Default is "./PhotoReceptorData.pkl".
    - ledDATA_path (str, optional): Path to the LED spectral data file. Default is "./IlluminationData.pkl".
    - acDATA (dict, optional): Dictionary mapping opsin names to their respective collecting areas. 
      Default values are provided for common opsins.
    - x_axis (str, optional): Key name for wavelength data in both opsin and LED files. Default is "x_axis".

    Returns:
    - numpy.ndarray: A 2D array where:
      - Rows correspond to selected opsins.
      - Columns correspond to selected LEDs.
      - Values represent isomerisation rates.

    Notes:
    - Opsin and LED spectra are normalized before processing.
    - Negative spectral values are clipped to zero.
    - An assertion checks that the wavelength range match between opsin and LED datasets.
    """


    h = 6.63 * 10**(-34)  # Planck constant in J.s
    c = 299792458         # Speed of light in m/s
    
    # Collecting areas for selected opsins
    Ac = np.array([acDATA[ac] for ac in selected_opsins])

    # Load opsin spectra data
    opsinsSpectrumsDATA = np.load(opsinDATA_path, allow_pickle=True)
    opsins = np.array([opsinsSpectrumsDATA[opsin_name] / opsinsSpectrumsDATA[opsin_name].max() for opsin_name in selected_opsins]) #Make sure it is normalized
    opsins[opsins < 0] = 0  # Set negative opsin values to 0
    
    # Load LED spectra data
    ledSpectrumDATA = np.load(ledDATA_path, allow_pickle=True)
    leds = np.array([ledSpectrumDATA[led] / ledSpectrumDATA[led].max() for led in selected_leds]) #Make sure it is normalized
    leds[leds < 0] = 0  # Set negative LED values to 0

    # Wavelengths (lambdas) for opsins and LEDs (they should be the same)
    lambdas_opsin = opsinsSpectrumsDATA[x_axis]
    lambdas_led = ledSpectrumDATA[x_axis]
    assert np.all(lambdas_opsin == lambdas_led), "Leds and Opsins spectrum don't match their lambdas"

    lambdas = lambdas_led
    dlambdas = np.diff(lambdas)
    dlambdas = np.hstack((np.mean(dlambdas[:3]), dlambdas))  # Ensure same length for lambdas and dlambdas

    leds = np.clip(leds / (np.sum(leds * dlambdas, axis=1)[:, None]), 0, None)   # Computing P(λ) / Integral(P(λ)dλ) and remove values under 0

    
    # Now we calculate the isomerisation without the loop.
    # Create a factor that doesn't depend on the wavelength (planck constant, speed of light, and collecting areas)
    normalization_factor = lambdas * dlambdas / (h * c) * 10**(-23)
    
    # Reshape for broadcasting and vectorized multiplication
    opsins_expanded = (opsins * Ac[:, None])  # Shape: (n_opsins, wavelengths)

    # Perform the element-wise multiplication across all wavelengths and then sum across the last axis (wavelengths)
    return np.sum(opsins_expanded[:, None, :] * leds[None, :, :] * normalization_factor, axis=2)


def led_solution_to_Ptot(led_dict, led_default_order=None):
    """
    Convert an LED power dictionary into a NumPy array with a fixed LED order. This is only used as a formating function for "get_isomerisations" or "plot_isomerisations"

    Parameters:
        led_dict (dict): Dictionary with LED colors as keys and power as values.
        led_order (list, optional): List defining the fixed order of LEDs.
                                    Defaults to ['Violet', 'Blue', 'Green', 'Yellow', 'Red'].

    Returns:
        ndarray: NumPy array of LED power values in the specified order.
    """
    if led_default_order is None:
        led_default_order = ['Violet', 'Blue', 'Green', 'Yellow', 'Red']
    
    return np.array([led_dict.get(led, 0) for led in led_default_order])

def get_isomerisations(
    Ptot_list,
    selected_LEDs,
    matrix = False,
    selected_opsins = ["Scones", "Mela", "Rods", "Mcones", "RedOpsin"],
    all_LEDs = ['Violet', 'Blue', 'Green', 'Yellow', 'Red'],
    colors = ['Violet', 'Blue', 'Green', 'orange', 'Red'],
    verbose = True,
    opsinDATA_path='./PhotoReceptorData.pkl',
    ledDATA_path='./IlluminationData.pkl',
    acDATA={'Scones': 0.2, 'Mcones': 0.2, 'RedOpsin': 0.002, 'Rods': 0.5, 'Mela': 0.2},
    x_axis='x_axis',
):
    
    """

    Computes isomerisation rates for selected opsins under specified LED illumination Power. If you just want a simple
    computation use this function. Here isomerisation_matrix will be computed everytime. If you want higher performance 
    computing consider imitating this function's behaviour (ex: computing lots of isomerisation rates with the same
    isomerisation matrix).  

    Parameters:
                    -------------- Main variables --------------

    - Ptot_list (list): List of power values for selected LED (list or numpy array)
    - selected_LEDs (list): List of selected LED names.
    - matrix (bool, optional): If True, returns the full isomerisation matrix; otherwise, returns total isomerisation per opsin. Default is False.
    
                    ------------ Adjusting variables ------------

    
    - selected_opsins (list, optional): List of opsins to consider. Default includes common opsins.
    - all_LEDs (list, optional): List of all possible LEDs. Default includes standard LED colors.
    - verbose (bool, optional): If True, prints detailed output. Default is True.
    
                    ---------- Very specific variables ----------
                    
    - opsinDATA_path (str, optional): Path to the opsin data file used in `get_isomerisation_matrix`.
    - ledDATA_path (str, optional): Path to the LED data file used in `get_isomerisation_matrix`.
    - acDATA (dict, optional): Dictionary of opsin-specific adaptation coefficients used in `get_isomerisation_matrix`.
    - x_axis (str, optional): Parameter controlling the x-axis representation in `get_isomerisation_matrix`.
    
                    ----------------------------------------------

    Returns:
    
    - numpy.ndarray: If matrix is True, returns a matrix of shape (len(selected_opsins), len(all_LEDs)), where each value represents the contribution of each LED to each opsin's isomerisation.
      If matrix is False, returns a 1D array of total isomerisation rates per opsin.
      The sum along axis=1 of the matrix result is equivalent to the total isomerisation output when matrix is False.

    Notes :
    - The function requires `get_isomerisation_matrix` and `led_solution_to_Ptot` to be defined.

    """
    
    assert isinstance(selected_LEDs, list), "selected_LEDs must be a list of LED names."
    assert all(isinstance(led, str) for led in selected_LEDs), "Each LED in selected_LEDs must be a string."
    assert len(Ptot_list) == len(selected_LEDs), "Ptot_list and selected_LEDs must have the same length."
    assert isinstance(matrix, bool), "matrix must be a boolean value."
    assert isinstance(selected_opsins, list), "selected_opsins must be a list."
    assert all(isinstance(opsin, str) for opsin in selected_opsins), "Each opsin in selected_opsins must be a string."
    assert isinstance(all_LEDs, list), "all_LEDs must be a list of LED names."
    assert all(isinstance(led, str) for led in all_LEDs), "Each LED in all_LEDs must be a string."
    assert isinstance(verbose, bool), "verbose must be a boolean value."
    assert all(led in all_LEDs for led in selected_LEDs), "You are passing a led that is not defined in the default leds. if this is not an error, please input a full all_LEDs list of leds"
    
    isomerisation_matrix = get_isomerisation_matrix(selected_opsins, all_LEDs, opsinDATA_path = opsinDATA_path, ledDATA_path=ledDATA_path, acDATA=acDATA, x_axis=x_axis) 

    Ptot = led_solution_to_Ptot({led: power for led, power in zip(selected_LEDs, Ptot_list)}, led_default_order = all_LEDs)
    if matrix:
        isomerisation = isomerisation_matrix * Ptot
        tot_isomerisation = isomerisation.sum(axis=1)
    else: 
        isomerisation = (isomerisation_matrix @ Ptot.reshape((-1,1))).reshape(-1)
        tot_isomerisation = isomerisation
    if verbose:
        print(f"\nPtot with {len(selected_LEDs)} LED Power Values (µW/cm²):")
        for i, led in enumerate(selected_LEDs):
            print(f"{led}: {Ptot[i]:.3f}")

        print("\nIsomerisation rate:")
        for i, opsin in enumerate(selected_opsins):
            print(f"{opsin}: {int(tot_isomerisation[i]):.1e}") 

    return isomerisation

def get_led_color(spectrum, x_axis, cmap_name="nipy_spectral"):
    cmap = matplotlib.colormaps[cmap_name]
    norm_x = (x_axis - x_axis.min()) / (x_axis.max() - x_axis.min())
    return tuple(np.round(np.average(np.array([cmap(n) for n in norm_x]), axis=0, weights=spectrum/spectrum.max())[:-1], decimals=3))


def plot_isomerisations(
    Ptot_list,
    selected_LEDs,
    selected_opsins=["Scones", "Mela", "Rods", "Mcones", "RedOpsin"],
    all_LEDs=None,
    colors={'Violet':'Violet', 'Blue':'Blue', 'Green':'Green', 'Yellow':'orange', 'Red':'Red', 'White':'grey'},
    verbose=True,
    opsinDATA_path="./PhotoReceptorData.pkl", 
    ledDATA_path="./IlluminationData.pkl", 
    acDATA={"Scones": 0.2, "Mcones": 0.2, "RedOpsin": 0.002, "Rods": 0.5, "Mela": 0.2}, 
    x_axis="x_axis"
):
    """
    Plots isomerisation rates for selected opsins under specified LED illumination power.
    
    This function visualizes how different LED power settings affect the isomerisation rates of various opsins.
    The input consists of LED power values (µW/cm²) and their corresponding LED names. Multiple LED configurations
    can be plotted by providing lists of power values and LED sets.
    
    Parameters:
                    -------------- Main variables --------------
    - Ptot_list (list of lists or numpy arrays): List containing power values (µW/cm²) for different LED settings. 
      Each entry should be a list or array corresponding to a specific illumination condition.
    - selected_LEDs (list of lists): List of lists containing the names of LEDs corresponding to each power set in `Ptot_list`.
    
                    ------------ Adjusting variables ------------
    
    - selected_opsins (list, optional): List of opsins to consider in the plot. Defaults to common opsins.
    - all_LEDs (list, optional): List of all available LEDs. Each `selected_LEDs` entry must be a subset of this list.
    - colors (dict, optional): Dict of colors for each LED in the plot. Must align with `all_LEDs`. Defaults to standard LED colors.
    - verbose (bool, optional): If True, prints LED power values and computed isomerisation rates. Default is True.
    
                    ---------- Very specific variables ----------
    
    - opsinDATA_path (str, optional): Path to the opsin data file used in `get_isomerisation_matrix`.
    - ledDATA_path (str, optional): Path to the LED data file used in `get_isomerisation_matrix`.
    - acDATA (dict, optional): Dictionary of opsin-specific adaptation coefficients used in `get_isomerisation_matrix`.
    - x_axis (str, optional): Parameter controlling the x-axis representation in `get_isomerisation_matrix`.
                    ----------------------------------------------
                    
    Returns:
    - matplotlib.figure.Figure: A bar plot where:
      - Each opsin is represented on the x-axis.
      - The y-axis shows the isomerisation rates.
      - Stacked bars represent contributions from different LEDs.
      - Different LED configurations are offset for clarity.

    Notes:
    - The function requires `get_isomerisations` to be defined.
    - `selected_LEDs` must be a subset of `all_LEDs`; otherwise, an assertion error will be raised.
    
    """
    if isinstance(Ptot_list, np.ndarray):
        Ptot_list = [Ptot_list]
    elif isinstance(Ptot_list, list):
        if not isinstance(Ptot_list[0], list) and not isinstance(Ptot_list[0], np.ndarray):
            Ptot_list = [np.array(Ptot_list)]
        else:
            Ptot_list = [np.array(p) if isinstance(p, list) else p for p in Ptot_list]
    else:
        raise TypeError("Ptot_list must be a numpy array, a list of numpy arrays, or a list of lists of numbers.")


    if not isinstance(selected_LEDs[0], list):
        selected_LEDs = [selected_LEDs]
        
    if not all_LEDs:
        all_LEDs = list(set([led for leds in selected_LEDs for led in leds ]))
        
    
    ledSpec = ledSpectrums(path = ledDATA_path, plot=False)
    for led in all_LEDs:
        if led not in colors.keys():
            colors[led] = get_led_color(ledSpec[led], ledSpec[x_axis])

            
    num_Ptot = len(Ptot_list)
    width = 0.2  # Width of each bar
    spacing = 0.1  # Space between grouped bars
    total_width = num_Ptot * width + (num_Ptot - 1) * spacing  # Total width of a grouped bar set
    dx_values = np.linspace(-total_width / 2 + width / 2, total_width / 2 - width / 2, num_Ptot)
    
    fig, ax = plt.subplots()

    for idx, (Ptot, dx) in enumerate(zip(Ptot_list, dx_values)):
        isomerisationMatrix_Ptot = get_isomerisations(
            Ptot,
            selected_LEDs[idx],
            matrix=True,
            selected_opsins=selected_opsins,
            all_LEDs = selected_LEDs[idx],
            verbose=verbose,
            opsinDATA_path=opsinDATA_path,
            ledDATA_path=ledDATA_path,
            acDATA=acDATA,
            x_axis=x_axis
        )
        
        positions = np.arange(isomerisationMatrix_Ptot.shape[0]) * max(1,len(Ptot_list)//2)
        accumulated_heights = np.zeros(isomerisationMatrix_Ptot.shape[0])  
        for i in range(isomerisationMatrix_Ptot.shape[1]):
            bars = ax.bar(
                positions + dx, 
                isomerisationMatrix_Ptot[:, i], 
                width, 
                bottom=accumulated_heights, 
                color=colors[selected_LEDs[idx][i]], 
                label=selected_LEDs[idx][i] if idx == 0 else None
            )
            accumulated_heights += isomerisationMatrix_Ptot[:, i]

        # Add value labels only at the top of the stacked bars
        for pos, height in zip(positions + dx, accumulated_heights):
            ax.text(
                pos, 
                height / 2, 
                f'{int(height):.1e}',  # Scientific notation format
                ha='center', 
                va='bottom', 
                fontsize=8,
                color='Black',
                rotation='vertical',
                fontweight='bold'
            )

    ax.set_ylabel('Isomerization')
    ax.set_title('Isomerization by Opsin and LED')
    ax.set_xticks(positions)
    ax.set_xticklabels(selected_opsins)
    ax.legend(title="LEDs")
    ax.grid(True, axis='y')
    plt.tight_layout()
    
    return fig


def interactive_Ptot_slider(
    selected_LEDs=None, 
    selected_opsins=None, 
    opsinDATA_path="./PhotoReceptorData.pkl", 
    ledDATA_path="./IlluminationData.pkl",
    x_axis="x_axis",
    acDATA=None,
    max_vals=None,
    def_vals=None,
    colors = None,
    FONT="Arial"
):
    
    """
    This function generates an interactive plot based on user input for various LED intensities and opsin data.

    Parameters:
    ----------
    selected_LEDs : list of str, optional
        A list of LEDs to be used for generating the plot. Each LED should correspond to an entry in the 'max_vals', 
        'def_vals', and 'colors' dictionaries. 
        Default is : ['Violet', 'Blue', 'Green', 'Yellow', 'Red'].

    selected_opsins : list of str, optional
        A list of opsins to consider in the calculations. These should be opsins included in the opsin spectrum file. 
        Default is : ["Scones", "Mela", "Rods", "Mcones", "RedOpsin"].

    opsinDATA_path : str, optional
        The file path to the opsin data (usually a pickle file). This file should contain necessary information 
        for opsin-specific parameters. 
        Default is : './PhotoReceptorData.pkl'.

    ledDATA_path : str, optional
        The file path to the LED data (usually a pickle file). This file should contain the illumination 
        properties for each LED. 
        Default : './IlluminationData.pkl'.

    acDATA : dict, optional
        A dictionary with opsin names as keys and their associated alpha coefficients (numeric values) as values. 
        Default values are provided for common opsins: {'Scones': 0.2, 'Mcones': 0.2, 'RedOpsin': 0.002, 'Rods': 0.5, 'Mela': 0.2}.

    max_vals : dict, optional
        A dictionary with each LED as a key and the maximum intensity value for that LED as the value. 
        Default is : {'Violet': 5, 'Blue': 5, 'Green': 5, 'Yellow': 5, 'Red': 500}.

    def_vals : dict, optional
        A dictionary with each LED as a key and the default intensity value for that LED as the value. 
        Default is : {'Violet': 1, 'Blue': 1, 'Green': 1, 'Yellow': 1, 'Red': 0}.

    colors : dict, optional
        A dictionary that maps each LED to a color. This is used for visual differentiation in the plot. 
        Default is : {'Violet': 'violet', 'Blue': 'blue', 'Green': 'green', 'Yellow': 'orange', 'Red': 'red'}.

    Returns:
    -------
    None : 
        The function will open a Tkinter window with interactive sliders for adjusting the LED intensities 
        and display a plot showing the isomerization values based on the selected LEDs and opsins.

    Usage:
    ------
    To call this function, you can either pass in your custom data dictionaries or use the default values.

    Example 1: Using default dictionaries
    -------------------
    your_function_name()

    Example 2: Providing custom dictionaries
    -------------------
    custom_max_vals = {'Violet': 10, 'Blue': 8, 'Green': 6, 'Yellow': 5, 'Red': 400}
    custom_def_vals = {'Violet': 3, 'Blue': 2, 'Green': 2, 'Yellow': 2, 'Red': 100}
    custom_colors = {'Violet': 'purple', 'Blue': 'lightblue', 'Green': 'lime', 'Yellow': 'yellow', 'Red': 'darkred'}

    your_function_name(
        selected_LEDs=['Violet', 'Blue', 'Green'],
        max_vals=custom_max_vals,
        def_vals=custom_def_vals,
        colors=custom_colors
    )

    Notes:
    -----
    - The 'opsinDATA_path' and 'ledDATA_path' should point to valid pickle files containing the required data for opsins and LEDs.
    - These should be dictionnaries with keydefining the names for LEDs and Opsins to use in all other dictionnaries and values
      being numpy arrays of the spectrums
    - All spectrum must be measured at the same wavelength so that opsins and leds share the same x axis
    - You may want to provid your own name for the x_axis as here it is called by default x_axis = "x_axis"
    - Ensure that all the LEDs listed in the 'selected_LEDs' list are also present in the 'max_vals', 'def_vals', 
      and 'colors' dictionaries. Otherwise, an assertion error will be raised.
    - The 'opsinDATA_path' and 'ledDATA_path' should point to valid pickle files containing the required data for opsins and LEDs.
    """
    
    # Default values if no arguments are provided
    if selected_LEDs is None:
        selected_LEDs = ['Violet', 'Blue', 'Green', 'Yellow', 'Red']
    
    if selected_opsins is None:
        selected_opsins = ["Scones", "Mela", "Rods", "Mcones", "RedOpsin"]
    
    if acDATA is None:
        acDATA = {"Scones": 0.2, "Mcones": 0.2, "RedOpsin": 0.002, "Rods": 0.5, "Mela": 0.2}
    
    if max_vals is None:
        max_vals = {'Violet': 5, 'Blue': 5, 'Green': 5, 'Yellow': 5, 'Red': 500}
    
    if def_vals is None:
        def_vals = {'Violet': 1, 'Blue': 1, 'Green': 1, 'Yellow': 1, 'Red': 0}

    if colors is None:
        colors = {'Violet' : 'violet', 'Blue' : 'blue', 'Green' : 'green', 'Yellow' : 'orange', 'Red' : 'red', "White":'grey'}

    for opsin in selected_opsins:
        if opsin not in acDATA:
            acDATA[opsin] = 0.2
            print(f"Opsin {opsin} has no defined Collecting Area defined in acDATA. Automatically setting to 0.2 as cones")
    
    ledSpec = ledSpectrums(path = ledDATA_path, plot=False)
    for led in selected_LEDs:
        if led not in max_vals:
            max_vals[led] = 20
            print(f"Led {led} not defined in 'max_vals'. Default value set at 5µW/cm2.")
        if led not in def_vals:
            def_vals[led] = 1
            print(f"Led not defined in 'def_vals'. Default value set to 1µw/cm2")
        if led not in colors.keys():
            rgb = 255*np.array(get_led_color(ledSpec[led], ledSpec[x_axis]) )
            colors[led] = f"#{int(rgb[0]):02X}{int(rgb[1]):02X}{int(rgb[2]):02X}"
            print(f"Led {led} color for plotting not defined. Plotting with automatic coloring.")
    assert all(led in max_vals and led in def_vals and led in colors for led in selected_LEDs), "Some LEDs in selected_LEDs are not detailed in max_vals, def_vals, or colors."
    assert all(opsin in acDATA for opsin in selected_opsins), "Some opsins are not defined in the acDATA dictionary."

    
    # Get isomerisation matrix
    isomerisation_matrix = get_isomerisation_matrix(
        selected_opsins, selected_LEDs, opsinDATA_path=opsinDATA_path, 
        ledDATA_path=ledDATA_path, acDATA=acDATA, x_axis=x_axis
    )

    # Function to update the plot based on user input
    def update_plot(Ptot_values):
        Ptot = np.array(Ptot_values)  # Convert the input values to a numpy array
        isomerisationMatrix = (isomerisation_matrix * Ptot.T)  # Compute the matrix
        
        fig = go.Figure()

        accumulated_heights = np.zeros(len(selected_opsins))  # Initialize accumulated heights to zero for each opsin
        cumulated_values = np.zeros(len(selected_opsins))  # Store cumulative values for annotations

        for i, led in enumerate(selected_LEDs):
            trace = go.Bar(
                x=selected_opsins,
                y=isomerisationMatrix[:, i],
                name=led,
                base=accumulated_heights,  # Stack the bars
                marker=dict(color=colors[led])  # Set the color for each LED
            )
            fig.add_trace(trace)
            accumulated_heights += isomerisationMatrix[:, i]  # Update the accumulated heights for the next LED
            cumulated_values += isomerisationMatrix[:, i]  # Update cumulative values

        fig.update_layout(
            barmode='stack',
            title="Isomerization of Opsins by LEDs",
            xaxis_title="Opsin",
            yaxis_title="Isomerization",
            legend_title="LEDs"
        )

        # Add cumulative values on top of each bar
        for i, value in enumerate(cumulated_values):
            fig.add_annotation(
                x=selected_opsins[i],
                y=value,
                text=f"{round(value)}",
                showarrow=True,
                arrowhead=2,
                ax=0,
                ay=-20,
                font=dict(size=10, color="black"),
                arrowcolor="black"
            )

        # Render the Plotly figure to a static image (PNG format)
        img_bytes = pio.to_image(fig, format='png')
        img = Image.open(io.BytesIO(img_bytes))
        img_tk = ImageTk.PhotoImage(img)

        # Update the image label with the new plot
        image_label.config(image=img_tk)
        image_label.image = img_tk  # Keep a reference to the image

    # Bind the slider value change to update the corresponding value label
    def update_value_label(slider, value_entry, i):
        value = Ptot_values[i].get()
        value_entry.delete(0, tk.END)
        value_entry.insert(0, f"{value:.3f}")
        update_plot([Ptot_values[j].get() for j in range(len(selected_LEDs))])

    def on_entry_change(value_entry, i):
        try:
            new_value = float(value_entry.get())
            if 0 <= new_value:
                Ptot_values[i].set(new_value)
                update_plot([Ptot_values[j].get() for j in range(len(selected_LEDs))])
            else:
                value_entry.delete(0, tk.END)
                value_entry.insert(0, f"{Ptot_values[i].get():.3f}")
        except ValueError:
            pass  # Ignore invalid input

    def reset_Ptot_values(event=None):
        for i, led in enumerate(selected_LEDs):
            val = def_vals[led]
            Ptot_values[i].set(val)
            value_entries[led].delete(0, tk.END)
            value_entries[led].insert(0, f"{val:.3f}")
        
        update_plot([Ptot_values[i].get() for i in range(len(selected_LEDs))])

        
    def print_values_on_close():
        # Print LED power values
        print("LED Power Values (µW/cm²):")
        for i, led in enumerate(selected_LEDs):
            print(f"{led}: {Ptot_values[i].get():.3f}")

        # Print Opsin Isomerization
        print("\nOpsin Isomerization Values:")
        Ptot = np.array([Ptot_values[i].get() for i in range(len(selected_LEDs))])
        isomerisationMatrix = (isomerisation_matrix@Ptot[:,None])  # Compute the matrix

        for i, opsin in enumerate(selected_opsins):
            print(f"{opsin}: {(isomerisationMatrix[i].astype(int))}")

        # Close the application
        root.destroy()
    
    
    # GUI setup using Tkinter
    root = tk.Tk()
    root.minsize(600, 650)
    root.maxsize(1000, 800)
    root.title("Interactive Ptot Slider")
    root.resizable(True, True)
    photo = tk.PhotoImage(file = "isomerisation_icon.png")
    root.wm_iconphoto(False, photo)

    # Create sliders for each LED in Ptot (each slider goes from 0 to max value for each LED)
    Ptot_values = [tk.DoubleVar(value=1.0) for _ in selected_LEDs]
    value_entries = {}  # List to store Entry widgets

    for i, led in enumerate(selected_LEDs):
        slider = ttk.Scale(root, from_=0, to=max_vals[led], orient="horizontal", variable=Ptot_values[i], length=300)
        slider.set(def_vals[led])
        slider.grid(row=i, column=1, padx=10, pady=10, sticky="nsew")
        
        label = ttk.Label(root, text=led, anchor="e", font=FONT + ' 13 bold')
        label.grid(row=i, column=0, padx=10, pady=10, sticky="nsew")
        
        frame = ttk.Frame(root)
        frame.grid(row=i, column=2, padx=5, pady=0, sticky="nsew")

        value_entry = ttk.Entry(frame, font=FONT + ' 10', width=8)
        value_entry.insert(0, f"{def_vals[led]:.3f}")
        value_entry.grid(row=0, column=0, padx=(0, 5), pady=10, sticky="nsew")

        unit_label = ttk.Label(frame, text="µW/cm²", font=FONT + ' 10')
        unit_label.grid(row=0, column=1, padx=(0, 0), pady=10, sticky="nsew")
        
        value_entries[led] = value_entry
        
        slider.bind("<Motion>", lambda event, slider=slider, value_entry=value_entry, i=i: update_value_label(slider, value_entry, i))
        value_entry.bind("<FocusOut>", lambda event, value_entry=value_entry, i=i: on_entry_change(value_entry, i))
        value_entry.bind("<Return>", lambda event, value_entry=value_entry, i=i: on_entry_change(value_entry, i))

    root.bind("<space>", reset_Ptot_values)
    root.protocol("WM_DELETE_WINDOW", print_values_on_close)  # Bind the close event

    image_label = ttk.Label(root, anchor='center')
    image_label.grid(row=len(selected_LEDs) + 1, columnspan=3, pady=20, sticky="nsew")

    # Configure grid resizing
    for i in range(len(selected_LEDs)):
        root.grid_rowconfigure(i, weight=1, minsize=33)
    root.grid_rowconfigure(len(selected_LEDs) + 1, weight=1, minsize=10)
    for i in range(3):
        root.grid_columnconfigure(i, weight=1, minsize=100)

    # Initialize with an empty plot
    update_plot([Ptot_values[i].get() for i in range(len(selected_LEDs))])

    # Run the GUI
    root.mainloop()

    
def solve_led_power_brut_force(matrix, target, high_bounds=None, low_bounds=None, initial_step=5, final_step=0.01):
    """
    Solve the LED power mix for given opsin isomerizations using a multi-step brute force approach.

    Parameters:
        matrix (ndarray): Matrix (n_opsins x n_leds) of isomerization efficiencies.
        target (ndarray): Vector (n_opsins,) of target isomerizations.
        high_bounds (ndarray): Upper bounds for each LED power.
        low_bounds (ndarray, optional): Lower bounds for each LED power (default: zeros).
        initial_step (float, optional): Initial step size for scanning the space.
        final_step (float, optional): Final step size for fine-tuning the solution.

    Returns:
        best_guess (ndarray): The LED power mix vector (n_leds,).
    """
    n_leds = matrix.shape[1]
    
    if high_bounds is None:
        high_bounds = 20*np.ones(len(n_leds))
    
    if low_bounds is None:
        low_bounds = np.zeros(n_leds)
    assert len(low_bounds) == len(high_bounds) == n_leds, f"Shapes are not consistent. Ensure bounds match the number of LEDs ({n_leds})."
    
    target = target.reshape((-1,1))
    
    def dist(guess):
        guess = guess.reshape((-1,1))
        return np.linalg.norm(matrix @ guess - target)
    
    def search_with_step(step, low_bounds, high_bounds, best_guess=None, best_error=float('inf')):
        """
        Recursive function to perform the search with a given step size.
        """
        # Ensure low_bounds and high_bounds are arrays of the correct shape
        low_bounds = np.asarray(low_bounds)
        high_bounds = np.asarray(high_bounds)

        # Create the search space
        search_space = [np.arange(low, high + step, step) for low, high in zip(low_bounds.flatten(), high_bounds.flatten())]

        total_tests = np.prod([len(s) for s in search_space])



        for guess in tqdm(itertools.product(*search_space), total=total_tests, desc=f"Searching with step {step}", leave=False):
            guess = np.array(guess)
            error = dist(guess)

            if error < best_error:
                best_error = error
                best_guess = guess

        return best_guess, best_error
    
    # First, scan the space with a large step size
    best_guess, best_error = search_with_step(initial_step, low_bounds, high_bounds)
    
    step = initial_step
    while step >= final_step:
        # Create a local search space around the current best guess
        local_low_bounds = np.maximum(low_bounds, best_guess - step)
        local_high_bounds = np.minimum(high_bounds, best_guess + step)
        step /= 10  # Gradually reduce the step size
        best_guess, best_error = search_with_step(step, local_low_bounds, local_high_bounds, best_guess, best_error)

    return best_guess
    
    
def solve_led_power_with_guess(matrix, target, initial_guess=None, low_bounds=None, high_bounds=None, weights=None, maxiter=1e7):
    """
    Solve the LED power mix for given opsin isomerizations with a starting condition.

    Parameters:
        A (ndarray): Matrix (n_opsins x n_leds) of isomerization efficiencies.
        I (ndarray): Vector (n_opsins,1) of target isomerizations.
        initial_guess (ndarray): Initial guess for the solution (n_leds,1).

    Returns:
        P (ndarray): The LED power mix vector (n_leds,).
    """
    import numpy as np
    import scipy.optimize
    
    n_leds = matrix.shape[1]
    
    target = np.array(target).reshape((-1, 1))
    
    if weights is not None:
        weights = weights.reshape((-1, 1))
    else:
        weights = np.zeros((matrix.shape[1], 1))
    
    def objective(guess):
        guess = guess.reshape((-1, 1))
        return np.linalg.norm(matrix @ guess - target) + np.linalg.norm(weights * guess)
    
    if low_bounds is None:
        low_bounds = np.zeros(n_leds)
    if high_bounds is None:
        high_bounds = [None for _ in range(n_leds)]
    if initial_guess is None:
        initial_guess = np.ones(n_leds)
    if weights is None:
        weights = np.zeros(n_leds)
    
    assert len(low_bounds) == len(high_bounds) == len(initial_guess) == len(weights) == n_leds, f"Shape are not consistent, check that low_bounds, high_bounds, initial_guess or weights are of size {n_leds}"    
    
    def track_optimization():
        history = []  # Store function values over iterations
        grad_history = []  # Store gradient norms

        def callback(xk):
            """Callback function to track function values and gradient norms during optimization."""
            history.append(objective(xk))
            grad = scipy.optimize.approx_fprime(xk, objective, epsilon=1e-8)
            grad_history.append(np.linalg.norm(grad))

        return history, grad_history, callback
    
    def optimize_with_tracking(initial_guess, low_bounds, high_bounds, maxiter=100):
        history, grad_history, callback = track_optimization()

        result = scipy.optimize.minimize(
            objective,
            initial_guess,
            bounds=[(low, high) for low, high in zip(low_bounds, high_bounds)],
            options={"maxiter": maxiter},
            tol=1e-3,
            callback=callback  # Track progress
        )

        converged =  np.diff(history)[-1] > -1e-3 # Check if function values stopped improving significantly

        if not result.success:
            if converged:
                print("Optimization Warning:", result.message)
                print("Optimization terminated abnormally but appears to have converged.")
            else:
                print("❌ Optimization Failed:", result.message)
                raise
        return result.x, history, grad_history
    
    opt_solution, fun_history, grad_history = optimize_with_tracking(initial_guess, low_bounds, high_bounds)
#     print("Function value history:", fun_history)
#     print("Gradient norm history:", grad_history)
    
    return opt_solution, fun_history, grad_history
    

def compute_balanced_weights(A, epsilon=1e-8):
    """
    Compute weights to balance the contribution of LEDs.

    Parameters:
        A (ndarray): Matrix (n_opsins x n_leds) of isomerization efficiencies.
        epsilon (float): Small constant to prevent division by zero.

    Returns:
        weights (ndarray): Weights for each LED (n_leds,).
    """
    column_norms = np.linalg.norm(A, axis=0)  # Compute L2 norms of columns
    weights = 1 / (column_norms + epsilon)   # Inverse proportionality
    return weights

def compute_VioletOverRed_weights(A, alpha=2):
    return alpha * np.linspace(0, 1, A.shape[1])

def compute_RedOverViolet_weights(A, alpha=2):
    return alpha * np.linspace(1, 0, A.shape[1])

def compute_priority_weights(manual_priorities, epsilon=1e-9):
    return 1 / ( np.array(manual_priorities) + epsilon)



def get_mix_color(isomerisation_target, 
    selected_LEDs = None, 
    mix_type = None, 
    opsinDATA_path = None, 
    ledDATA_path=None, 
    acDATA=None, 
    low_bounds = None,
    high_bounds = None,
    initial_guess = None,
    x_axis="x_axis"):
    
    """
    Computes the LED power mix required to achieve a target isomerisation profile. This function is a higher level function to solve_led_power_with_guess for convenience but it removes some freedom from user.

    This function determines the optimal power levels for a set of LEDs to match a given target 
    isomerisation rate for selected opsins. It allows for different mixing strategies to prioritize 
    certain LED contributions.

    Parameters:
    - isomerisation_target (dict): Dictionary mapping opsin names to their target isomerisation rates.
    - selected_LEDs (list, optional): List of LED names to be used in the mix. Defaults to standard LED colors.
    - mix_type (str, list, or numpy.ndarray, optional): Strategy for LED weighting. Options include:
      - 'balance': Balances contributions across LEDs.
      - 'violet_over_red': Prioritizes violet LEDs.
      - 'red_over_violet': Prioritizes red LEDs.
      - Custom priority vector (list or array) of the same length as `selected_LEDs`.
      
              ------------ tuning variables ------------ 
              (used to tune solce_led_poser_with_guess)
              
    - low_bounds (list, optional): Lower bounds for LED power values. Defaults to zero for all LEDs.
    - high_bounds (list, optional): Upper bounds for LED power values. Defaults to no constraint.
    - initial_guess (list, optional): Initial guess for LED power values. Defaults to ones for all LEDs.
      
              ------------ adjusting variables ------------ 
                  (Used to get the isomersiation_matrix)
                  
    - opsinDATA_path (str, optional): Path to the opsin spectral data file. Default is "./PhotoReceptorData.pkl".
    - ledDATA_path (str, optional): Path to the LED spectral data file. Default is "./IlluminationData.pkl".
    - acDATA (dict, optional): Dictionary mapping opsin names to their collecting areas. 
      Defaults to standard values.
    - x_axis (str, optional): Key name for wavelength data in the spectral files. Default is "x_axis".


    Returns:
    - numpy.ndarray or None: Array of computed LED power values (µW/cm²) for each LED in `selected_LEDs`.
      If the optimization fails, returns None.

    Notes:
    - The function uses `get_isomerisation_matrix` to compute the isomerisation rates.
    - `solve_led_power_with_guess` is used to solve for LED power values based on the selected strategy.
    - If an invalid `mix_type` is provided, a warning is printed, and all weights are set to zero.
    - If solving fails, an error message is printed, and None is returned.
    """
    
    
    selected_opsins = list(isomerisation_target.keys())
    
    if selected_LEDs is None:
        selected_LEDs = ['Violet', 'Blue', 'Green', 'Yellow', 'Red']
    if opsinDATA_path is None:
        opsinDATA_path = "./PhotoReceptorData.pkl"
    if ledDATA_path is None:
        ledDATA_path="./IlluminationData.pkl"
    if acDATA is None:
        acDATA = {"Scones": 0.2, "Mcones": 0.2, "RedOpsin": 0.002, "Rods": 0.5, "Mela": 0.2}
    
    isomerisation_matrix = get_isomerisation_matrix(selected_opsins, selected_LEDs, opsinDATA_path = opsinDATA_path, ledDATA_path = ledDATA_path, acDATA = acDATA, x_axis='x_axis')
    
    n_leds = len(selected_LEDs)
    
    if low_bounds is None:
        low_bounds = np.zeros(n_leds)
    
    if high_bounds is None:
        high_bounds = [None for _ in range(n_leds)]

    if initial_guess is None:
        initial_guess = np.ones(n_leds)
    
    if mix_type is None: 
        weights = np.zeros((n_leds,1))
        
    elif mix_type == 'balance':
        weights = compute_balanced_weights(isomerisation_matrix).reshape((-1,1))
        
    elif mix_type == 'violet_over_red':
        weights = compute_VioletOverRed_weights(isomerisation_matrix).reshape((-1,1))
    
    elif mix_type == 'red_over_violet':
        weights = compute_RedOverViolet_weights(isomerisation_matrix).reshape((-1,1))
    
    elif type(mix_type) is np.array or type(mix_type) is list:
        assert len(mix_type) == n_leds, "You are trying to pass a priority vector for leds to turn on but the priority vector is not the same length than the number of leds!"
        weights = compute_priority_weights(mix_type).reshape((-1,1))
    else:
        print("Unknown mix_type. Weights all put to zeros. Available mix_types are 'balance', 'fewViolet', 'fewRed', Priority by inputing the a priority vector of same length than n_leds.")
        low_bounds = np.zeros((n_leds,1))

    try:
        P,_,_ = solve_led_power_with_guess(isomerisation_matrix, np.array(list(isomerisation_target.values())), initial_guess = initial_guess, low_bounds=low_bounds, high_bounds=high_bounds, weights=weights)
        
        print("LED Power Mix:\n" + "\n".join(f"{selected_LEDs[i]}: {P[i]:.2f} µW/cm²" for i in range(n_leds)))
    except ValueError as e:
        print("Error:", e)
        P=None
    return P

def get_list_from_vec(file_path, col_index=2):
    """
    Reads a .vec file and extracts a specified column into two lists:
    - A list of unique values without consecutive repetition.
    - A list of counts of consecutive occurrences of each unique value.

    Parameters:
    ----------
    file_path : str
        Path to the .vec file.
    col_index : int, optional
        The 0-based index of the column to extract. Default is 2 (third column).

    Returns:
    -------
    unique_values : list
        A list of unique values without consecutive repetition.
    counts : list
        A list containing the number of consecutive occurrences of each unique value.
    """
    data = np.loadtxt(file_path, skiprows=1, usecols=[col_index], dtype=int)

    unique_values = []
    series = []

    for i in range(len(data)):
        if data[i] == data[i - 1]:
            series[-1] += 1
        else:
            unique_values.append(data[i])
            series.append(1)

    return unique_values, series


def load_obj(name):
    """
        Generic function to load a bin obj with pickle protocol

    Input :
        - name (str) : path to where the obj is
    Output :
        - (python object) : loaded object
        
    Possible mistakes :
        - Wrong path 
    """
    
    if os.path.dirname(os.path.normpath(name)) != '':
        os.makedirs(os.path.dirname(os.path.normpath(name)), exist_ok=True)
    else:
        name = os.path.join(os.getcwd(),os.path.normpath(name))
    if name[-4:]!='.pkl':
        name += '.pkl'
    with open(os.path.normpath(name), 'rb') as f:
        return pickle.load(f)
    
def save_obj(obj, name ):
    """
        Generic function to save an obj with pickle protocol

    Input :
        - obj (python var) : object to be saved in binary format
        - name (str) : path to where the obj shoud be saved

    Possible mistakes :
        - Permissions denied, restart notebook from an admin shell
        - Folders aren't callable, change your folders
    """
    
    if os.path.dirname(os.path.normpath(name)) != '':
        os.makedirs(os.path.dirname(os.path.normpath(name)), exist_ok=True)
    else:
        name = os.path.join(os.getcwd(),os.path.normpath(name))
    
    if name[-4:]!='.pkl':
        name += '.pkl'
    with open( os.path.normpath(name), 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def charge_calibration(path, calibration_date=r'20250225', 
                       all_LEDs = ['Violet', 'Blue', 'Green', 'Yellow', 'Red'],
                      verbose = True):

    print('\nC O R R E C T I O N \n')
    
    correction = {'Red':None,
                  'Yellow':None,
                  'Green':None,
                  'Blue':None,
                  'Violet':None, 
                 }
    for i in range(len(all_LEDs)):
        temp_color_name = all_LEDs[i]
        temp_corr = input('Enter the measured power of the {} LED at 5V (mW): '.format(temp_color_name) )
        try :
            temp_corr = float(temp_corr)
            print('The {} LED power function will be corrected.'.format(temp_color_name) )
            correction[temp_color_name]=temp_corr
        except :
            print('No correction will be applied to the {} LED.'.format(temp_color_name))
            pass

    print('\nF I L T E R S \n')
    
    filters = {'Red':1,
                'Yellow':1,
                'Green':1,
                'Blue':1,
                'Violet':1, 
                 }

    for i in range(len(all_LEDs)):
        temp_color_name = all_LEDs[i]
        temp_filter = input('Enter the filter transmittance value applied to the {} LED: '.format(temp_color_name))
        try :
            temp_filter = float(temp_filter)
            print('The {} LED will be filtered.'.format(temp_color_name))
            filters[temp_color_name]=temp_filter
        except :
            print('No filter will be applied to the {} LED.'.format(temp_color_name))
            pass
    
    all_LEDs = np.flip(all_LEDs)
    wb = pxl.load_workbook(path)
    ws = wb[calibration_date]
    #col_number = 5
    #colors = ['625 nm','530 nm','490 nm', '415 nm', '385 nm']
    
    fiber_to_mea_red = {'Red':1,
                        'Yellow':1,
                        'Green':1,
                        'Blue':1,
                        'Violet':1, 
                 }
    np.ones(len(all_LEDs))
    
    calibrations = {'voltages' : np.array([ws.cell(row=i, column=1).value for i in range(4,21)])}
    for col_i in range(0,len(all_LEDs)):
        color_name = all_LEDs[col_i]
        temp_ratio = ws.cell(row=23, column=2+col_i).value/ws.cell(row=20, column=2+col_i).value
        fiber_to_mea_red[color_name]=temp_ratio 
        
        calibrations[color_name]=np.array([ws.cell(row=i, column=2+col_i).value for i in range(4,21)])

        
        if correction[color_name]!=None :
            calibrations[color_name]=calibrations[color_name]*correction[color_name]/calibrations[color_name][-1]
    
        calibrations[color_name]=calibrations[color_name]*filters[color_name]*fiber_to_mea_red[color_name]
        print(calibrations[color_name])
    min_pow=0
    ind=1
    old_pow=calibrations[all_LEDs[0]][ind]
    while min_pow==0 :
        for col_name in all_LEDs :
            temp_pow=calibrations[col_name][ind]
            min_pow = min(temp_pow,old_pow)
            old_pow=temp_pow
        ind+=1
    if ind>=1 :
        calibrations['voltages']=np.concatenate(([0],calibrations['voltages'][ind:]))
        for col_name in all_LEDs :
            calibrations[col_name]=np.concatenate(([0],calibrations[col_name][ind:]))


    if verbose :
        for col_name in all_LEDs :
        
            plt.plot((calibrations['voltages']),calibrations[col_name], label=col_name)
            plt.yscale('log')
            plt.xlabel('Tension (V)')
            plt.ylabel('Power (µW/cm²)')
            plt.title('{} LED'.format(col_name))
            plt.legend()
            plt.show(block=False)
    
    return calibrations, fiber_to_mea_red

def get_voltages(Ptot, calibration, all_LEDs = ['Violet', 'Blue', 'Green', 'Yellow', 'Red'], verbose=False):

    voltages = calibration['voltages']
    driving_tension = []
    
    for col_i in range(len(all_LEDs)) :
        col_name = all_LEDs[col_i]
        temp_P = float(Ptot[col_i])
                        
        # Return 0 voltage if the power is 0
        if temp_P == 0:
            driving_tension.append(0)
            if verbose:
                print(f"{col_name}: Power is 0, so voltage is 0 V.")
            continue

        
        calibration_list = calibration[col_name]
        if temp_P>calibration_list[-1]:
            raise ValueError('The given power value is too high. Try putting a lower value.')
    
        if temp_P in calibration_list :
            index=np.where(calibration_list==temp_P)
            res_voltage=voltages[index][0]
        
        else : 
            low_ind=0
            high_ind=len(calibration_list)-1
            
            for i in range(0,len(calibration_list)):
                if calibration_list[i]<temp_P and calibration_list[i]>calibration_list[low_ind]:
                    # print(calibration_list[i])
                    low_ind=i
                if calibration_list[i]>temp_P and calibration_list[i]<calibration_list[high_ind]:
                    high_ind=i
                #print(voltages[low_ind],voltages[high_ind])
            
            res_voltage = voltages[low_ind] + ( voltages[high_ind] - voltages[low_ind] ) * ( temp_P - calibration_list[low_ind] ) / ( calibration_list[high_ind] - calibration_list[low_ind] )
            driving_tension.append(res_voltage)
            
            if verbose:
                print('{}   :    {} V'.format(col_name, res_voltage) )
            
    driving_tension = np.array(driving_tension)
    return driving_tension

